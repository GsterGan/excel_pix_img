from PIL import Image
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import colors,Font,Color,PatternFill,Border,Side
import winsound
import os
import re

def setColour(img,name):
    wb = Workbook()
    filename = name + '.xlsx'
    ws = wb.active
    bd = Side(style='thick', color="FFFFFF")
    print(len(img),len(img[0]))
    for row in range(1,len(img[0])):
        for col in range(1,len(img)):
            location = str(get_column_letter(col)) + str(row)
            #ws2[location] = '█'#█●
            ws.column_dimensions[get_column_letter(col)].width= 2.5
            ws[location].fill = PatternFill("solid",fgColor=img[col][row])
            ws[location].border = Border(left=bd, top=bd, right=bd, bottom=bd)
            #print(img[row][col])
    wb.save(filename = filename)
    print('done')

def getPic(path,size):
    heximg = []
    img = Image.open(path)
    x,y = img.size
    img = img.resize((size,int(y*size/x)))
    #img.show()
    img = np.array(img)
    for line in img:
        linehex = []
        for pix in line:
            linehex.append(toHex(pix[0],pix[1],pix[2]))
        heximg.append(linehex)
    heximg = np.transpose(heximg)
    print(len(heximg),len(heximg[0]))
    return heximg

def toHex(r, g, b):
    color = 'FF'
    color += str(hex(r)).replace('x','0')[-2:]
    color += str(hex(g)).replace('x','0')[-2:]
    color += str(hex(b)).replace('x','0')[-2:]
    return color

def walk_path(path,size):
    walk = os.walk(path)
    for root,dirs,files in walk:
        for file in files:
            if file.endswith(".jpg"):
                heximg = getPic(os.path.join(root,file),size)
                setColour(heximg,file.split(".")[0])
#    winsound.Beep(1000,1000)

def one_deal(path,size):
    heximg = getPic(path,size)
    setColour(heximg,path.split('.')[0])
#    winsound.Beep(1000,1000)

if __name__ == "__main__":
    path = 'E:/excel_pix_img/out/lib.jpg'
    #walk_path()
    one_deal(path,100)

